import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "outputs" / "vendor_invoice_autopilot_workbook.xlsx"


def main():
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_csv_sheet(workbook, "vendors", ROOT / "data" / "vendors.csv")
    add_csv_sheet(workbook, "demand_forecasts", ROOT / "data" / "demand_forecasts.csv")
    add_csv_sheet(workbook, "inventory_position", ROOT / "data" / "inventory_position.csv")
    add_csv_sheet(workbook, "supplier_rules", ROOT / "data" / "supplier_rules.csv")
    add_csv_sheet(workbook, "purchase_orders", ROOT / "data" / "purchase_orders.csv")
    add_csv_sheet(workbook, "invoice_reviews", ROOT / "data" / "invoice_reviews_template.csv")
    add_csv_sheet(workbook, "demo_reviews_backup", ROOT / "out" / "invoice_reviews.csv")

    OUTPUT_PATH.parent.mkdir(exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH.relative_to(ROOT)}")


def add_csv_sheet(workbook, title, csv_path):
    sheet = workbook.create_sheet(title)
    with csv_path.open(newline="", encoding="utf-8") as file:
        reader = csv.reader(file)
        for row in reader:
            sheet.append(row)

    header_fill = PatternFill("solid", fgColor="E7F5EE")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        width = min(max(max_length + 2, 12), 48)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


if __name__ == "__main__":
    main()
